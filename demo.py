from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

file_name = 'blood_donation.xlsx'
valid_blood_types = {'A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'}

# Initialize the workbook if it does not exist
if not os.path.exists(file_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Donors'
    sheet.append(['ID', 'Name', 'Age', 'Blood Type', 'Last Donation Date'])
    workbook.save(file_name)

def get_next_id():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Donors']
    
    max_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            max_id = max(max_id, row[0])
    
    return max_id + 1

def validate_donor(data):
    if 'age' in data and (data['age'] < 18 or data['age'] > 60):
        return False, "Age must be between 18-60."
    
    if 'blood_type' in data:
        data['blood_type'] = data['blood_type'].upper()
        if data['blood_type'] not in valid_blood_types:
            return False, "Invalid blood type."
    
    if 'last_donation_date' in data:
        try:
            donation_date = datetime.strptime(data['last_donation_date'], '%Y-%m-%d')
            if donation_date > datetime.today():
                return False, "Last donation date cannot be in the future."
        except ValueError:
            return False, "Invalid date format. Use YYYY-MM-DD."
    
    return True, ""

@app.route('/add_donor', methods=['POST'])
def add_donor():
    data = request.json
    is_valid, message = validate_donor(data)
    if not is_valid:
        return jsonify({"message": message}), 400
    
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Donors']
        
        new_id = get_next_id()
        sheet.append([new_id, data['name'], data['age'], data['blood_type'], data['last_donation_date']])
        workbook.save(file_name)
        
        return jsonify({"message": "Donor added successfully!"})
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

@app.route('/view_donors', methods=['GET'])
def view_donors():
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Donors']
        
        donors = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            donors.append(row)
        
        return jsonify(donors)
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

def validate_donor(data):
    # Implement your validation logic here if needed
    return True, "Valid data"

@app.route('/update_donor/<int:id>', methods=['PUT'])
def update_donor(id):
    data = request.json
    
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Donors']
        
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == id:
                if 'name' in data and data['name'] is not None:
                    row[1].value = data['name']
                if 'age' in data and data['age'] is not None:
                    row[2].value = data['age']
                if 'blood_type' in data and data['blood_type'] is not None:
                    row[3].value = data['blood_type'].upper()
                if 'last_donation_date' in data and data['last_donation_date'] is not None:
                    row[4].value = data['last_donation_date']
                workbook.save(file_name)
                return jsonify({"message": "Donor updated successfully!"})
        
        return jsonify({"message": "Donor not found!"}), 404
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

@app.route('/delete_donor/<int:id>', methods=['DELETE'])
def delete_donor(id):
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Donors']
        
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == id:
                sheet.delete_rows(row[0].row)
                workbook.save(file_name)
                return jsonify({"message": "Donor deleted successfully!"})
        
        return jsonify({"message": "Donor not found!"}), 404
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

@app.route('/count_donors', methods=['GET'])
def count_donors():
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Donors']
        
        count = sheet.max_row - 1
        
        return jsonify({"count": count})
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

@app.route('/search_donors/<blood_type>', methods=['GET'])
def search_donors(blood_type):
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Donors']
        
        donors = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == blood_type:
                donors.append(row)
        
        if not donors:
            return jsonify({"message": "No donors found with the specified blood type."}), 404
        
        return jsonify(donors)
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True)
